"""
Excel Generator - Generates realistic Excel files for import testing.
"""
import os
from datetime import datetime, timedelta
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd

from src.seeders.master_seeder import MasterSeeder
from src.seeders.machine_seeder import MachineSeeder


class ExcelGenerator:
    """
    Generates Excel test files for import testing:
    - Master data Excel files per class
    - Machine data Excel with user and log sheets
    """
    
    def __init__(self, db_session, output_dir: str = 'tests/datasets'):
        """
        Initialize the Excel generator.
        
        Args:
            db_session: SQLAlchemy database session
            output_dir: Directory to save generated files
        """
        self.db = db_session
        self.output_dir = output_dir
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
    
    def generate_all(self) -> Dict[str, str]:
        """
        Generate all Excel test files.
        
        Returns:
            Dictionary mapping file types to paths
        """
        result = {}
        
        # Generate master data files for each class
        for class_id, class_name in [('CLASS_9A', '9A'), ('CLASS_9B', '9B'), ('CLASS_9C', '9C')]:
            path = self.generate_master_excel(class_id, class_name)
            result[f'master_{class_name}'] = path
        
        # Generate machine data file
        path = self.generate_machine_excel()
        result['machine'] = path
        
        return result
    
    def generate_master_excel(self, class_id: str, class_name: str) -> str:
        """
        Generate master data Excel file for a specific class.
        
        Format:
        Row 1-2: School header (merged cells)
        Row 3: "DAFTAR HADIR SISWA"
        Row 4: "Kls : 9A"
        Row 5: "Wali Kelas : Budi Santoso"
        Row 6: Empty
        Row 7: Table header | NO | NO. INDUK | NAMA | ... |
        Row 8+: Student data
        """
        wb = Workbook()
        ws = wb.active
        ws.title = 'Daftar Hadir'
        
        # Get class data
        all_students = MasterSeeder.get_all_students()
        class_students = [s for s in all_students if s['class_id'] == class_id]
        
        # Get wali kelas name
        wali_kelas_map = {
            'CLASS_9A': 'Budi Santoso',
            'CLASS_9B': 'Siti Aminah',
            'CLASS_9C': 'Ahmad Hidayat'
        }
        wali_name = wali_kelas_map.get(class_id, 'Unknown')
        
        # Header rows
        ws.merge_cells('A1:F2')
        ws['A1'] = 'SMP NEGERI 1 CONTOH'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A3'] = 'DAFTAR HADIR SISWA'
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = f'Kls : {class_name}'
        ws['A5'] = f'Wali Kelas : {wali_name}'
        
        # Table header row
        header_row = 7
        headers = ['NO', 'NO. INDUK', 'NAMA', 'TELEPON ORTU']
        
        # Add date columns for current month
        today = datetime.now()
        for day in range(1, 32):
            try:
                date = today.replace(day=day)
                headers.append(str(day))
            except ValueError:
                break
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Student data rows
        for idx, student in enumerate(class_students, 1):
            row = header_row + idx
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=student['nis'])
            ws.cell(row=row, column=3, value=student['name'])
            ws.cell(row=row, column=4, value=student['parent_phone'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = f'Daftar_Absen_Siswa_{class_name}.xlsx'
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
    
    def generate_machine_excel(self) -> str:
        """
        Generate machine data Excel with user and log sheets.
        
        Sheet 1: "Att. Stat. Report"
        Row 1: "Statistical Report"
        Row 2: "Generated: 2024-12-16"
        Row 3: Empty
        Row 4: Header | ID | Name | Department | Total Logs |
        Row 5+: User data
        
        Sheet 2: "Att.log report"
        Row 1: Header | ID | Name | DateTime | Device | Status |
        Row 2+: Log data
        """
        wb = Workbook()
        
        # Sheet 1: Statistical Report
        ws1 = wb.active
        ws1.title = 'Att. Stat. Report'
        
        ws1['A1'] = 'Statistical Report'
        ws1['A1'].font = Font(bold=True, size=14)
        
        ws1['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d")}'
        
        # Header row
        stat_headers = ['ID', 'Name', 'Department', 'Total Logs']
        for col, header in enumerate(stat_headers, 1):
            cell = ws1.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Get machine users data
        all_students = MasterSeeder.get_all_students()
        typo_variations = MachineSeeder.get_typo_variations()
        unmapped_users = MachineSeeder.UNMAPPED_USERS
        
        row = 5
        user_id = 195
        
        # Perfect match users
        typo_names = [t[0] for t in typo_variations]
        for student in all_students:
            if student['name'] not in typo_names:
                ws1.cell(row=row, column=1, value=str(user_id))
                ws1.cell(row=row, column=2, value=student['name'])
                ws1.cell(row=row, column=3, value='Student')
                ws1.cell(row=row, column=4, value=28)  # ~2 scans/day * 14 days
                row += 1
                user_id += 1
        
        # Typo variation users
        for _, machine_name in typo_variations:
            ws1.cell(row=row, column=1, value=str(user_id))
            ws1.cell(row=row, column=2, value=machine_name)
            ws1.cell(row=row, column=3, value='Student')
            ws1.cell(row=row, column=4, value=28)
            row += 1
            user_id += 1
        
        # Unmapped users
        for staff in unmapped_users:
            ws1.cell(row=row, column=1, value=str(user_id))
            ws1.cell(row=row, column=2, value=staff['name'])
            ws1.cell(row=row, column=3, value=staff['department'])
            ws1.cell(row=row, column=4, value=14)  # ~1 scan/day
            row += 1
            user_id += 1
        
        # Sheet 2: Attendance Log
        ws2 = wb.create_sheet('Att.log report')
        
        log_headers = ['ID', 'Name', 'DateTime', 'Device', 'Status']
        for col, header in enumerate(log_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Generate sample log entries
        row = 2
        today = datetime.now()
        user_id = 195
        
        # Add sample logs for first few students
        for i, student in enumerate(all_students[:5]):
            for day_offset in range(3):
                log_date = today - timedelta(days=day_offset)
                
                # Check-in
                checkin = log_date.replace(hour=7, minute=30 + i)
                ws2.cell(row=row, column=1, value=str(user_id + i))
                ws2.cell(row=row, column=2, value=student['name'])
                ws2.cell(row=row, column=3, value=checkin.strftime('%Y-%m-%d %H:%M:%S'))
                ws2.cell(row=row, column=4, value='MAIN_GATE')
                ws2.cell(row=row, column=5, value='Check In')
                row += 1
                
                # Check-out
                checkout = log_date.replace(hour=14, minute=30 + i)
                ws2.cell(row=row, column=1, value=str(user_id + i))
                ws2.cell(row=row, column=2, value=student['name'])
                ws2.cell(row=row, column=3, value=checkout.strftime('%Y-%m-%d %H:%M:%S'))
                ws2.cell(row=row, column=4, value='MAIN_GATE')
                ws2.cell(row=row, column=5, value='Check Out')
                row += 1
        
        # Auto-adjust column widths for both sheets
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        filename = f'Mapping_Mesin_{datetime.now().strftime("%b%Y")}.xlsx'
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        
        return filepath
